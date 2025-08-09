import pandas as pd
import pymssql
import os
from openpyxl import Workbook
import configparser
import mysql.connector
from cryptography.fernet import Fernet
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import datetime


###########################################################
# Codigo para retirar o relatorio e enviar para o usuario # 
########################################################### 


def load_key_from_env(env_var):
    key = os.getenv(env_var)
    if key is None:
        raise ValueError(f"Chave não encontrada na variável de ambiente: {env_var}")
    return key.encode()

web_server1_mysql = 'web_server1_mysql.ini'
key_env_var = 'CHAVE_web_server1'

def decrypt_config_file(input_file, key):
    cipher_suite = Fernet(key)
    with open(input_file, 'rb') as f:
        f.read(4) 
        encrypted_config_data = f.read()
    config_data = cipher_suite.decrypt(encrypted_config_data).decode()
    return config_data


current_dir = os.path.dirname(os.path.abspath(__file__))
web_server1_mysql_path = os.path.join(current_dir, web_server1_mysql)


key = load_key_from_env(key_env_var)
config_data = decrypt_config_file(web_server1_mysql_path, key)

if config_data is not None:
    config = configparser.ConfigParser()
    config.read_string(config_data)

# conexao
destino_host = config['mysql_destino_db']['host']
destino_database = config['mysql_destino_db']['database']
destino_user = config['mysql_destino_db']['user']
destino_password = config['mysql_destino_db']['password']
destino_port = config['mysql_destino_db']['port']

destino_db = mysql.connector.connect(
    host=destino_host,
    user=destino_user,
    port=destino_port,
    password=destino_password,
    database=destino_database
)

cursor_destino = destino_db.cursor()

#Querys
query1 = """SELECT 
    CASE
         WHEN usu_sitcvs = 'A' THEN 'ATIVO'
         WHEN usu_sitcvs = 'I' THEN 'CANCELADO'
    END AS 'Status Contrato',
    ...
    ;"""
query2 = """SELECT customer_id as ID_Empresa,
	   ...
)"""

#Conexão com SQL Server
origem_host = config['sqlserver_origem_db']['host']
origem_database = config['sqlserver_origem_db']['database']
origem_user = config['sqlserver_origem_db']['user']
origem_password = config['sqlserver_origem_db']['password']


server = origem_host
database = origem_database
username = origem_user
password = origem_password


conn = pymssql.connect(server=server, user=username, password=password, database=database)

#Query SQL Server
cursor = conn.cursor()

query3 = """
SELECT usu_numctr 
...
"""


#Tratamento
df1 = pd.read_sql(query1, destino_db)
df2 = pd.read_sql(query2, destino_db)
df3 = pd.read_sql(query3, conn)

wb1 = Workbook()
ws1 = wb1.active

wb2 = Workbook()
ws2 = wb2.active

wb3 = Workbook()
ws3 = wb3.active

ws1.append(df1.columns.tolist())
for _, row in df1.iterrows():
    ws1.append(row.tolist())

ws2.append(df2.columns.tolist())
for _, row in df2.iterrows():
    ws2.append(row.tolist())

ws3.append(df3.columns.tolist())
for _, row in df3.iterrows():
    ws3.append(row.tolist())


#Data para montar arquivo e envio de e-mail
horario_atual = datetime.datetime.now()
data_formatada = horario_atual.strftime("%d-%m-%Y")
data_arquivo = horario_atual.strftime("%d-%m-%Y")

arquivo1 = f"Arquivo_para_Validar_{data_arquivo}.xlsx"
arquivo2 = f"Inconsistentes_{data_arquivo}.xlsx"
arquivo3 = f"projeto_ad_Automatizada_{data_arquivo}.xlsx"

wb1.save(arquivo1)
wb2.save(arquivo2)
wb3.save(arquivo3)

destino_db.close()
conn.close()

#Envio de e-mail
sender_email = "xxxxxxx" #remetente
receiver_email = ["destinatarios.gmail.com"] #destinatario
subject = f"Relatório projeto_ad - {data_formatada}"
body = """
Prezados(as),

Segue em anexo, os arquivos referente a projeto_ad.

....
"""

smtp_server = "smtp.gmail.com"
smtp_port = 587
smtp_username = "xxxx xxxx"  # google app user
smtp_password = "xxx xxx xxx" #app passworld 

msg = MIMEMultipart()
msg['From'] = sender_email
msg['To'] = ", ".join(receiver_email)
msg['Subject'] = subject

#Anexar arquivos
with open(arquivo1, "rb") as attachment:
    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename= {arquivo1}")
    msg.attach(part)

with open(arquivo2, "rb") as attachment:
    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename= {arquivo2}")
    msg.attach(part)

with open(arquivo3, "rb") as attachment:
    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename= {arquivo3}")
    msg.attach(part)    

msg.attach(MIMEText(body, "plain"))

with smtplib.SMTP(smtp_server, smtp_port) as server:
    server.starttls()
    server.login(smtp_username, smtp_password)
    server.sendmail(sender_email, receiver_email, msg.as_string())

#Remover os arquivos apos enviar
os.remove(arquivo1)
os.remove(arquivo2)
os.remove(arquivo3)